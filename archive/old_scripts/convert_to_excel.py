"""
Chuyển CSV sang Excel để xử lý phông chữ tiếng Việt
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

print("🔄 Chuyển CSV sang Excel...")

# Load CSV
csv_file = "AUTO_LABELED_FOR_REVIEW_20251229_015030.csv"
df = pd.read_csv(csv_file, encoding='utf-8')

print(f"✅ Loaded: {len(df)} rows")

# Save to Excel
excel_file = "AUTO_LABELED_FOR_REVIEW_20251229_015030.xlsx"
df.to_excel(excel_file, index=False, engine='openpyxl')

print(f"✅ Saved to Excel: {excel_file}")

# Format Excel
print("🎨 Formatting Excel...")

wb = load_workbook(excel_file)
ws = wb.active

# Set font cho toàn bộ sheet
font = Font(name='Arial', size=11)
alignment = Alignment(wrap_text=True, vertical='top')

for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    for cell in row:
        cell.font = font
        cell.alignment = alignment

# Set column widths
ws.column_dimensions['A'].width = 80  # training_text
ws.column_dimensions['B'].width = 10  # label
ws.column_dimensions['C'].width = 15  # confidence
ws.column_dimensions['D'].width = 25  # pattern
ws.column_dimensions['E'].width = 40  # note
ws.column_dimensions['F'].width = 20  # source

# Bold header
header_font = Font(name='Arial', size=11, bold=True)
for cell in ws[1]:
    cell.font = header_font

# Save
wb.save(excel_file)

print(f"✅ Formatted Excel saved: {excel_file}")

# Tạo các file Excel riêng theo confidence
for conf in ['HIGH', 'MEDIUM', 'LOW']:
    csv_conf = f"AUTO_LABELED_{conf}_20251229_015030.csv"
    excel_conf = f"AUTO_LABELED_{conf}_20251229_015030.xlsx"
    
    try:
        df_conf = pd.read_csv(csv_conf, encoding='utf-8')
        df_conf.to_excel(excel_conf, index=False, engine='openpyxl')
        
        # Format
        wb_conf = load_workbook(excel_conf)
        ws_conf = wb_conf.active
        
        for row in ws_conf.iter_rows(min_row=1, max_row=ws_conf.max_row, min_col=1, max_col=ws_conf.max_column):
            for cell in row:
                cell.font = font
                cell.alignment = alignment
        
        ws_conf.column_dimensions['A'].width = 80
        ws_conf.column_dimensions['B'].width = 10
        ws_conf.column_dimensions['C'].width = 15
        ws_conf.column_dimensions['D'].width = 25
        ws_conf.column_dimensions['E'].width = 40
        ws_conf.column_dimensions['F'].width = 20
        
        for cell in ws_conf[1]:
            cell.font = header_font
        
        wb_conf.save(excel_conf)
        print(f"✅ {conf}: {excel_conf} ({len(df_conf)} rows)")
    except:
        pass

print("\n✅ HOÀN THÀNH!")
print("Các file Excel đã được tạo với phông chữ Arial, wrap text, và column width phù hợp")
