"""
Cắt bỏ post title dài, chỉ giữ lại comment sau </s>
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
import re

print("✂️ Cắt bỏ post title dài...")

# Load data
df = pd.read_csv("AUTO_LABELED_FOR_REVIEW_20251229_015030.csv", encoding='utf-8-sig')
print(f"✅ Loaded: {len(df)} rows")

# Function để cắt title dài
def remove_long_title(text):
    """
    Nếu có </s> và phần trước </s> dài > 100 chars → Chỉ lấy comment sau </s>
    """
    if '</s>' not in text:
        return text
    
    parts = text.split('</s>')
    if len(parts) < 2:
        return text
    
    title = parts[0].strip()
    comment = parts[-1].strip()
    
    # Nếu title quá dài (>100 chars) → Chỉ lấy comment
    if len(title) > 100:
        return comment
    
    # Nếu title ngắn → Giữ nguyên
    return text

# Apply
print("✂️ Processing...")
df['training_text_original'] = df['training_text']
df['training_text'] = df['training_text'].apply(remove_long_title)

# Đếm số rows đã cắt
cut_count = (df['training_text'] != df['training_text_original']).sum()
print(f"✅ Đã cắt title dài: {cut_count} rows")

# Reorder columns
cols = ['training_text', 'label', 'confidence', 'pattern', 'note', 'source', 'training_text_original']
df = df[cols]

# Save to Excel
print("\n💾 Saving to Excel...")
excel_file = "AUTO_LABELED_SHORT_TITLE.xlsx"

wb = Workbook()
ws = wb.active
ws.title = "Short Title"

# Header
headers = ['training_text', 'label', 'confidence', 'pattern', 'note', 'source']
for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.font = Font(name='Arial', size=11, bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center')

# Data
for row_idx, row in enumerate(df[headers].itertuples(index=False), 2):
    for col_idx, value in enumerate(row, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=str(value))
        cell.font = Font(name='Arial', size=11)
        cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    if row_idx % 1000 == 0:
        print(f"  Progress: {row_idx}/{len(df)}")

# Set column widths
ws.column_dimensions['A'].width = 80   # training_text (ngắn hơn)
ws.column_dimensions['B'].width = 10   # label
ws.column_dimensions['C'].width = 15   # confidence
ws.column_dimensions['D'].width = 30   # pattern
ws.column_dimensions['E'].width = 40   # note
ws.column_dimensions['F'].width = 25   # source

wb.save(excel_file)
print(f"✅ Saved: {excel_file}")

# Tạo file HIGH confidence
print("\n✂️ Creating HIGH confidence (short title)...")
df_high = df[df['confidence'] == 'high']
excel_high = "AUTO_LABELED_HIGH_SHORT.xlsx"

wb_high = Workbook()
ws_high = wb_high.active
ws_high.title = "High Confidence"

for col_idx, header in enumerate(headers, 1):
    cell = ws_high.cell(row=1, column=col_idx, value=header)
    cell.font = Font(name='Arial', size=11, bold=True)

for row_idx, row in enumerate(df_high[headers].itertuples(index=False), 2):
    for col_idx, value in enumerate(row, 1):
        cell = ws_high.cell(row=row_idx, column=col_idx, value=str(value))
        cell.font = Font(name='Arial', size=11)
        cell.alignment = Alignment(wrap_text=True, vertical='top')

ws_high.column_dimensions['A'].width = 80
ws_high.column_dimensions['B'].width = 10
ws_high.column_dimensions['C'].width = 15
ws_high.column_dimensions['D'].width = 30
ws_high.column_dimensions['E'].width = 40
ws_high.column_dimensions['F'].width = 25

wb_high.save(excel_high)
print(f"✅ Saved: {excel_high} ({len(df_high)} rows)")

# Tạo file MEDIUM confidence (sample)
print("\n✂️ Creating MEDIUM confidence (short title, sample)...")
df_medium = df[df['confidence'] == 'medium'].head(1000)
excel_medium = "AUTO_LABELED_MEDIUM_SHORT.xlsx"

wb_medium = Workbook()
ws_medium = wb_medium.active
ws_medium.title = "Medium Confidence"

for col_idx, header in enumerate(headers, 1):
    cell = ws_medium.cell(row=1, column=col_idx, value=header)
    cell.font = Font(name='Arial', size=11, bold=True)

for row_idx, row in enumerate(df_medium[headers].itertuples(index=False), 2):
    for col_idx, value in enumerate(row, 1):
        cell = ws_medium.cell(row=row_idx, column=col_idx, value=str(value))
        cell.font = Font(name='Arial', size=11)
        cell.alignment = Alignment(wrap_text=True, vertical='top')

ws_medium.column_dimensions['A'].width = 80
ws_medium.column_dimensions['B'].width = 10
ws_medium.column_dimensions['C'].width = 15
ws_medium.column_dimensions['D'].width = 30
ws_medium.column_dimensions['E'].width = 40
ws_medium.column_dimensions['F'].width = 25

wb_medium.save(excel_medium)
print(f"✅ Saved: {excel_medium} ({len(df_medium)} rows)")

# Statistics
print("\n" + "="*80)
print("📊 THỐNG KÊ")
print("="*80)
print(f"Tổng số rows: {len(df)}")
print(f"Đã cắt title dài: {cut_count} rows ({cut_count/len(df)*100:.1f}%)")
print(f"\nFiles đã tạo:")
print(f"  1. {excel_file} - Tất cả {len(df)} rows")
print(f"  2. {excel_high} - HIGH confidence {len(df_high)} rows")
print(f"  3. {excel_medium} - MEDIUM confidence (sample 1000 rows)")

# Show examples
print("\n📝 Ví dụ đã cắt:")
cut_examples = df[df['training_text'] != df['training_text_original']].head(3)
for idx, row in cut_examples.iterrows():
    print(f"\n[{idx+1}]")
    print(f"  Trước: {row['training_text_original'][:150]}...")
    print(f"  Sau:   {row['training_text'][:150]}...")

print("\n✅ HOÀN THÀNH!")
