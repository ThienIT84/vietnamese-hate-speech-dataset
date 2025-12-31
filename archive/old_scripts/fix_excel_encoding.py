"""
Fix Excel encoding - Đảm bảo tiếng Việt hiển thị đúng
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

print("🔧 Fixing Excel encoding...")

# Load CSV với encoding đúng
csv_file = "AUTO_LABELED_FOR_REVIEW_20251229_015030.csv"
df = pd.read_csv(csv_file, encoding='utf-8-sig')  # utf-8-sig để xử lý BOM

print(f"✅ Loaded: {len(df)} rows")
print(f"📝 Sample text: {df.iloc[0]['training_text'][:100]}")

# Tạo Excel mới thủ công
excel_file = "AUTO_LABELED_REVIEW_FIXED.xlsx"

wb = Workbook()
ws = wb.active
ws.title = "Auto Labeled Data"

# Write header
headers = list(df.columns)
for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.font = Font(name='Arial', size=11, bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center')

# Write data
for row_idx, row in enumerate(df.itertuples(index=False), 2):
    for col_idx, value in enumerate(row, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=str(value))
        cell.font = Font(name='Arial', size=11)
        cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    if row_idx % 1000 == 0:
        print(f"  Progress: {row_idx}/{len(df)}")

# Set column widths
ws.column_dimensions['A'].width = 100  # training_text
ws.column_dimensions['B'].width = 10   # label
ws.column_dimensions['C'].width = 15   # confidence
ws.column_dimensions['D'].width = 30   # pattern
ws.column_dimensions['E'].width = 50   # note
ws.column_dimensions['F'].width = 25   # source

# Save
wb.save(excel_file)
print(f"✅ Saved: {excel_file}")

# Tạo file HIGH confidence
print("\n🔧 Creating HIGH confidence file...")
df_high = df[df['confidence'] == 'high']
excel_high = "AUTO_LABELED_HIGH_FIXED.xlsx"

wb_high = Workbook()
ws_high = wb_high.active
ws_high.title = "High Confidence"

# Header
for col_idx, header in enumerate(headers, 1):
    cell = ws_high.cell(row=1, column=col_idx, value=header)
    cell.font = Font(name='Arial', size=11, bold=True)

# Data
for row_idx, row in enumerate(df_high.itertuples(index=False), 2):
    for col_idx, value in enumerate(row, 1):
        cell = ws_high.cell(row=row_idx, column=col_idx, value=str(value))
        cell.font = Font(name='Arial', size=11)
        cell.alignment = Alignment(wrap_text=True, vertical='top')

ws_high.column_dimensions['A'].width = 100
ws_high.column_dimensions['B'].width = 10
ws_high.column_dimensions['C'].width = 15
ws_high.column_dimensions['D'].width = 30
ws_high.column_dimensions['E'].width = 50
ws_high.column_dimensions['F'].width = 25

wb_high.save(excel_high)
print(f"✅ Saved: {excel_high} ({len(df_high)} rows)")

# Tạo file MEDIUM confidence
print("\n🔧 Creating MEDIUM confidence file...")
df_medium = df[df['confidence'] == 'medium']
excel_medium = "AUTO_LABELED_MEDIUM_FIXED.xlsx"

wb_medium = Workbook()
ws_medium = wb_medium.active
ws_medium.title = "Medium Confidence"

# Header
for col_idx, header in enumerate(headers, 1):
    cell = ws_medium.cell(row=1, column=col_idx, value=header)
    cell.font = Font(name='Arial', size=11, bold=True)

# Data (chỉ lấy 1000 rows đầu để file không quá nặng)
for row_idx, row in enumerate(df_medium.head(1000).itertuples(index=False), 2):
    for col_idx, value in enumerate(row, 1):
        cell = ws_medium.cell(row=row_idx, column=col_idx, value=str(value))
        cell.font = Font(name='Arial', size=11)
        cell.alignment = Alignment(wrap_text=True, vertical='top')

ws_medium.column_dimensions['A'].width = 100
ws_medium.column_dimensions['B'].width = 10
ws_medium.column_dimensions['C'].width = 15
ws_medium.column_dimensions['D'].width = 30
ws_medium.column_dimensions['E'].width = 50
ws_medium.column_dimensions['F'].width = 25

wb_medium.save(excel_medium)
print(f"✅ Saved: {excel_medium} (1000/{len(df_medium)} rows - sample)")

print("\n✅ HOÀN THÀNH!")
print("\nFiles đã tạo:")
print(f"  1. {excel_file} - Tất cả {len(df)} rows")
print(f"  2. {excel_high} - HIGH confidence {len(df_high)} rows")
print(f"  3. {excel_medium} - MEDIUM confidence (sample 1000 rows)")
print("\nHãy thử mở bằng Microsoft Excel hoặc LibreOffice Calc")
